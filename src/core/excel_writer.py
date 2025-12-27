from ..config import base_config, excel_config
from .tariff import Tariff
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List


class ExcelWriter:
    '''
    Класс для создания и заполнения Excel файла с тарифами
    '''
    
    def __init__(self):
        self.filename = Path(base_config.OUTPUT_DIR) / base_config.OUTPUT_FILENAME

        self.wb = Workbook()

        self.sheets = {}
        self.current_row = {}

    
    def create_sheet(self):
        '''
        Создание нового листа в книге
        '''
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        ws = self.wb.create_sheet(excel_config.SHEET_NAME) # ????????????????
        self.sheets[excel_config.SHEET_NAME] = ws # ????????????????
        self.current_row[excel_config.SHEET_NAME] = 1

        return ws
    

    def write_headers(self):
        '''
        Записывает заголовки таблицы
        '''
        ws = self.sheets[excel_config.SHEET_NAME]
        current_row = self.current_row[excel_config.SHEET_NAME]
        headers = excel_config.COLUMNS
        
        for col, header in enumerate (headers, 1) :
            cell = ws.cell(row=current_row, column = col, value = header)
            self._get_header_style(cell, current_row, ws)

        self.current_row[excel_config.SHEET_NAME] = current_row + 1

    
    def _get_header_style(self, cell, current_row, ws): 
        '''
        Добавляет стили в заголовок таблицы
        '''
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) 
        ws.column_dimensions['A'].width = 60
        ws.row_dimensions[current_row].height = 30
        for col_letter in ['B', 'C', 'D']:
            ws.column_dimensions[col_letter].width = 15
        border_style = Side(style='thin', color='000000')
        border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    
    def write_tarriffs(self, tariffs: List[Tariff]):
        '''
        Записывает тарифы в таблицу
        '''
        ws = self.sheets[excel_config.SHEET_NAME]
        current_row = self.current_row[excel_config.SHEET_NAME]

        for tariff in tariffs:
            row_data = tariff.to_excel_row()
            for column, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=column, value=value)
                self._get_cell_style(cell)
        
            current_row += 1
        self.current_row[excel_config.SHEET_NAME] = current_row


    def _get_cell_style(self, cell):
        '''
        Добавляет стили ячейкам таблицы
        '''
        border_style = Side(style='thin', color='000000')
        border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        cell.alignment = alignment


    def save(self):
        '''
        Сохраняет книгу в файл
        '''
        try:
            self.filename.parent.mkdir(parents=True, exist_ok=True)
            self.wb.save(str(self.filename))
            return str(self.filename)
            
        except Exception:
            raise
